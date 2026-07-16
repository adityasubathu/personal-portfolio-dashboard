import csv
import io
import re
from collections import defaultdict
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from sqlalchemy import and_, delete, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.allocation_target import AllocationTarget, AssetClassTarget
from app.models.holding import Holding
from app.models.instrument import Instrument
from app.models.mf_breakdown import AmfiMarketCap, EquityCategoryOverride, MfSchemeBreakdown
from app.models.nav_history import NavHistory
from app.models.price_history import PriceHistory
from app.models.trade import Trade
from app.time_util import now_ist

BREAKDOWN_DIR = Path("data/mf_portfolio_breakdown")

_AMFI_DATE_RE = re.compile(
    r"AverageMarketCapitalization(\d{1,2})(\w{3})(\d{4})",
    re.IGNORECASE,
)
_MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

# ETFs whose entire equity portfolio is a single market-cap category.
# Fallback cap category for holdings in these ETFs that AMFI doesn't recognise.
# AMFI always takes priority; this only applies when AMFI returns Unclassified Equity.
ETF_CAP_OVERRIDE: dict[str, str] = {
    "INF204KB14I2": "Large Cap",   # NIFTYBEES
    "INF732E01045": "Large Cap",   # JUNIORBEES (Nifty Next 50)
    "INF769K01IC9": "Mid Cap",     # MIDCAPETF
    "INF0R8F01141": "Small Cap",   # SML100CASE
}

# Funds whose equity holdings are entirely foreign; bypasses AMFI market-cap lookup.
FOREIGN_FUND_ISINS: set[str] = {
    "INF247L01AP3",   # MON100 / Nasdaq 100
}

# Individual company names (normalised, lowercase substrings) that are always foreign equity,
# regardless of which fund holds them.
FOREIGN_COMPANY_SUBSTRINGS: set[str] = {
    "alphabet", "amazon", "apple", "meta platforms", "microsoft",
}

# Commodity ETFs whose entire value maps to a single non-equity category.
# These bypass MF breakdown CSV lookup entirely.
COMMODITY_ETF_CATEGORY: dict[str, str] = {
    "INF109KC1Y56": "Silver",   # SILVERIETF
    "INF204KB17I5": "Gold",     # GOLDBEES
    "INF0R8F01042": "Gold",     # GOLDCASE
}

_BRACKET_RE = re.compile(r"[\[(].*?[\])]")
_GLUED_DOT = re.compile(r"\.(?=[a-z])")
_ABBREV_MAP = [
    (re.compile(r"\bltd\.?(?=\W|$)", re.IGNORECASE), "limited"),
    (re.compile(r"\bcorpn?\.?(?=\W|$)", re.IGNORECASE), "corporation"),
    (re.compile(r"\bpvt\.?(?=\W|$)", re.IGNORECASE), "private"),
    (re.compile(r"\bco\.?(?=\W|$)", re.IGNORECASE), "company"),
]
_SUFFIX_RE = re.compile(
    r"\b(limited|limted|company|corporation|ordinary\s+shares|private)\b",
    re.IGNORECASE,
)
_TRAILING_JUNK = re.compile(r"[\s.*\-]+$")
_MULTI_SPACE = re.compile(r"\s+")

_NAME_ALIASES: dict[str, str] = {
    "m.r.f.": "mrf",
    "m r f": "mrf",
}


def normalize_company_name(name: str) -> str:
    s = name.lower().strip()
    s = _BRACKET_RE.sub("", s)
    s = _GLUED_DOT.sub(". ", s)
    for pat, expansion in _ABBREV_MAP:
        s = pat.sub(expansion, s)
    s = _SUFFIX_RE.sub("", s)
    s = s.replace("(", " ").replace(")", " ").replace(".", " ")
    s = _TRAILING_JUNK.sub("", s)
    s = _MULTI_SPACE.sub(" ", s).strip()
    for alias, canonical in _NAME_ALIASES.items():
        s = s.replace(alias, canonical)
    return s


def _find_amfi_xlsx() -> Path | None:
    if not BREAKDOWN_DIR.exists():
        return None
    candidates = list(BREAKDOWN_DIR.glob("AverageMarketCapitalization*.xlsx"))
    if not candidates:
        return None
    candidates.sort(key=lambda p: _parse_amfi_date(p.name) or date.min, reverse=True)
    return candidates[0]


def _parse_amfi_date(filename: str) -> date | None:
    m = _AMFI_DATE_RE.search(filename)
    if not m:
        return None
    day, mon_str, year = int(m.group(1)), m.group(2).lower(), int(m.group(3))
    month = _MONTH_MAP.get(mon_str[:3])
    if not month:
        return None
    return date(year, month, day)


def _load_sector_master() -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    """Returns ({isin: sector}, {normalized_name: sector}, {nse_symbol: sector})."""
    path = BREAKDOWN_DIR / "sector_master.csv"
    if not path.exists():
        return {}, {}, {}
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        text = path.read_text(encoding="latin-1")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return {}, {}, {}
    reader.fieldnames = [f.strip() for f in reader.fieldnames]
    isin_sector: dict[str, str] = {}
    name_sector: dict[str, str] = {}
    symbol_sector: dict[str, str] = {}
    for row in reader:
        isin = (row.get("ISIN Code") or "").strip()
        sector = (row.get("Industry") or "").strip()
        company = (row.get("Company Name") or "").strip()
        symbol = (row.get("Symbol") or "").strip()
        if not sector:
            continue
        if isin:
            isin_sector[isin] = sector
        if company:
            name_sector[normalize_company_name(company)] = sector
        if symbol:
            symbol_sector[symbol] = sector
    return isin_sector, name_sector, symbol_sector


def _write_company_master(rows: list[dict]) -> None:
    master_path = BREAKDOWN_DIR / "company_master.csv"
    BREAKDOWN_DIR.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "isin", "canonical_name", "primary_ticker", "nse_symbol", "bse_symbol",
        "msei_symbol", "exchanges", "mcap_category", "sector", "aliases",
    ]
    cat_order = {"Large Cap": 0, "Mid Cap": 1, "Small Cap": 2}
    sorted_rows = sorted(rows, key=lambda r: (cat_order.get(r["categorization"], 9), r["name_normalized"]))
    with master_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in sorted_rows:
            writer.writerow({
                "isin": r.get("isin") or "",
                "canonical_name": r.get("company_name") or "",
                "primary_ticker": r.get("primary_ticker") or "",
                "nse_symbol": r.get("nse_symbol") or "",
                "bse_symbol": r.get("bse_symbol") or "",
                "msei_symbol": r.get("msei_symbol") or "",
                "exchanges": r.get("exchanges") or "",
                "mcap_category": r.get("categorization") or "",
                "sector": r.get("sector") or "",
                "aliases": r.get("aliases") or "",
            })


async def sync_amfi_market_cap(db: AsyncSession, on_progress=None) -> dict:
    xlsx_path = _find_amfi_xlsx()
    if not xlsx_path:
        return {"error": "No AverageMarketCapitalization*.xlsx found in data/mf_portfolio_breakdown/"}

    file_date = _parse_amfi_date(xlsx_path.name)
    stale_warning = None
    if file_date:
        age_days = (date.today() - file_date).days
        if age_days > 180:
            stale_warning = f"Classification file is {age_days} days old ({file_date:%d %b %Y}). Consider updating from AMFI website."

    # Preserve user-edited aliases from existing company_master.csv
    master_path = BREAKDOWN_DIR / "company_master.csv"
    preserved_aliases: dict[str, str] = {}
    if master_path.exists():
        try:
            text = master_path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            text = master_path.read_text(encoding="latin-1")
        r = csv.DictReader(io.StringIO(text))
        for mrow in r:
            isin_key = (mrow.get("isin") or "").strip()
            alias_val = (mrow.get("aliases") or "").strip()
            if isin_key and alias_val:
                preserved_aliases[isin_key] = alias_val

    if on_progress:
        await on_progress(f"Loading AMFI data from {xlsx_path.name}…")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_to_insert: list[dict] = []
    counts = {"Large Cap": 0, "Mid Cap": 0, "Small Cap": 0}

    _DASH = {"-", "–", "—"}

    for row in ws.iter_rows(min_row=3, values_only=True):
        company = str(row[1] or "").strip()
        isin = str(row[2] or "").strip() or None
        bse_raw = str(row[3] or "").strip()
        nse_raw = str(row[5] or "").strip()
        msei_raw = str(row[7] or "").strip()
        cat = str(row[10] or "").strip()
        if not company or cat not in counts:
            continue

        bse_sym = bse_raw if bse_raw and bse_raw not in _DASH else None
        nse_sym = nse_raw if nse_raw and nse_raw not in _DASH else None
        msei_sym = msei_raw if msei_raw and msei_raw not in _DASH else None

        primary_ticker = nse_sym or bse_sym or msei_sym

        exchanges_parts: list[str] = []
        if nse_sym:
            exchanges_parts.append("NSE")
        if bse_sym:
            exchanges_parts.append("BSE")
        if msei_sym:
            exchanges_parts.append("MSEI")
        exchanges = ",".join(exchanges_parts) or None

        counts[cat] += 1
        rows_to_insert.append({
            "company_name": company,
            "isin": isin,
            "bse_symbol": bse_sym,
            "nse_symbol": nse_sym,
            "msei_symbol": msei_sym,
            "primary_ticker": primary_ticker,
            "exchanges": exchanges,
            "aliases": preserved_aliases.get(isin or "") or None,
            "categorization": cat,
            "sector": None,
            "name_normalized": normalize_company_name(company),
            "updated_at": now_ist(),
        })

    wb.close()

    if on_progress:
        await on_progress(
            f"Parsed {len(rows_to_insert)} companies "
            f"(Large: {counts['Large Cap']}, Mid: {counts['Mid Cap']}, Small: {counts['Small Cap']})"
        )

    # Enrich with sector: ISIN → NSE symbol → normalized name
    isin_sector, name_sector, symbol_sector = _load_sector_master()
    sectors_loaded = len(isin_sector)
    if isin_sector or name_sector or symbol_sector:
        for row in rows_to_insert:
            sec = isin_sector.get(row["isin"] or "")
            if not sec and row["nse_symbol"]:
                sec = symbol_sector.get(row["nse_symbol"])
            if not sec:
                sec = name_sector.get(row["name_normalized"] or "")
            if sec:
                row["sector"] = sec
        if on_progress:
            await on_progress(f"Sector enrichment: {sectors_loaded} ISIN mappings applied")

    await db.execute(delete(AmfiMarketCap))
    if rows_to_insert:
        await db.execute(AmfiMarketCap.__table__.insert(), rows_to_insert)
    await db.flush()

    _write_company_master(rows_to_insert)
    if on_progress:
        await on_progress("Company master CSV updated")

    result = {
        "rows_loaded": len(rows_to_insert),
        "large": counts["Large Cap"],
        "mid": counts["Mid Cap"],
        "small": counts["Small Cap"],
        "sectors_loaded": sectors_loaded,
        "file": xlsx_path.name,
    }
    if file_date:
        result["file_date"] = file_date.strftime("%d %b %Y")
    if stale_warning:
        result["stale_warning"] = stale_warning
    return result


def _resolve_equity_category(
    name: str,
    alias_to_isin: dict[str, str],
    name_to_isin: dict[str, str],
    isin_to_mcap: dict[str, str],
    amfi_by_name: dict[str, str],
) -> str:
    norm = normalize_company_name(name)
    for lookup in (alias_to_isin, name_to_isin):
        isin = lookup.get(norm)
        if isin and isin in isin_to_mcap:
            return isin_to_mcap[isin]
    cat = amfi_by_name.get(norm)
    if cat:
        return cat
    best_r, best_cat = 0.0, None
    for amfi_norm, amfi_cat in amfi_by_name.items():
        r = SequenceMatcher(None, norm, amfi_norm).ratio()
        if r > best_r:
            best_r, best_cat = r, amfi_cat
    if best_r >= 0.85 and best_cat:
        return best_cat
    return "Unclassified Equity"


def _resolve_equity_sector(
    name: str,
    alias_to_isin: dict[str, str],
    name_to_isin: dict[str, str],
    isin_to_sector: dict[str, str],
    name_to_sector: dict[str, str],
) -> str | None:
    norm = normalize_company_name(name)
    for lookup in (alias_to_isin, name_to_isin):
        isin = lookup.get(norm)
        if isin and isin in isin_to_sector:
            return isin_to_sector[isin]
    sec = name_to_sector.get(norm)
    if sec:
        return sec
    best_r, best_s = 0.0, None
    for amfi_norm, amfi_sec in name_to_sector.items():
        r = SequenceMatcher(None, norm, amfi_norm).ratio()
        if r > best_r:
            best_r, best_s = r, amfi_sec
    if best_r >= 0.85:
        return best_s
    return None


_MF_DEBT_RE = re.compile(r"\b(liquid|money\s+market|savings\s+fund|low\s+duration)\b", re.IGNORECASE)
_REIT_RE = re.compile(r"\b(reit|real\s+estate\s+trust)\b", re.IGNORECASE)


def _classify_type(holding_type: str, name: str) -> str:
    t = holding_type.strip()
    if t.startswith("Bond"):
        return "Debt"
    if t.startswith("Cash") or t == "Cash":
        return "Cash"
    if t.startswith("Mutual Fund") and (_MF_DEBT_RE.search(t) or _MF_DEBT_RE.search(name)):
        return "Debt"
    return "Other"


def _parse_holdings_pct(s: str) -> float | None:
    s = s.strip().rstrip("%").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _sector_for_type(holding_type: str, name: str) -> str | None:
    t = holding_type.strip()
    if t.startswith("Bond"):
        return "Fixed Income"
    if t.startswith("Cash") or t == "Cash":
        return "Liquid / Money Market"
    if t.startswith("Mutual Fund") and (_MF_DEBT_RE.search(t) or _MF_DEBT_RE.search(name)):
        return "Liquid / Money Market"
    return None


async def ingest_scheme_csvs(db: AsyncSession, on_progress=None) -> dict:
    amfi_all = (await db.execute(select(AmfiMarketCap))).scalars().all()

    alias_to_isin: dict[str, str] = {}
    name_to_isin: dict[str, str] = {}
    isin_to_mcap: dict[str, str] = {}
    isin_to_sector: dict[str, str] = {}
    amfi_by_name: dict[str, str] = {}
    amfi_name_sector: dict[str, str] = {}

    for a in amfi_all:
        amfi_by_name[a.name_normalized] = a.categorization
        if a.sector:
            amfi_name_sector[a.name_normalized] = a.sector
        if a.isin:
            isin_to_mcap[a.isin] = a.categorization
            name_to_isin[a.name_normalized] = a.isin
            if a.sector:
                isin_to_sector[a.isin] = a.sector
            if a.aliases:
                for alias in a.aliases.split("|"):
                    alias_norm = normalize_company_name(alias.strip())
                    if alias_norm:
                        alias_to_isin[alias_norm] = a.isin

    overrides: dict[str, str] = {
        o.name_normalized: o.category
        for o in (await db.execute(select(EquityCategoryOverride))).scalars().all()
    }

    held_funds = (await db.execute(
        select(Instrument)
        .join(Holding, Holding.instrument_id == Instrument.id)
        .where(Instrument.instrument_type.in_(("MF", "ETF")))
    )).scalars().all()
    held_isins = {i.isin for i in held_funds if i.isin}

    # Match "liquid" at a word start (no trailing boundary — catches LIQUIDCASE, LIQUIDBEES, etc.)
    _DEBT_KEYWORDS = re.compile(r"\b(debt|liquid)", re.IGNORECASE)
    _ARBITRAGE_RE = re.compile(r"\barbitrage\b", re.IGNORECASE)
    # No trailing \b — catches compound names like GoldCase, Silvercase, etc.
    _GOLD_RE = re.compile(r"\bgold", re.IGNORECASE)
    _SILVER_RE = re.compile(r"\bsilver", re.IGNORECASE)
    debt_fund_isins: set[str] = set()
    arbitrage_fund_isins: set[str] = set()
    gold_fund_isins: set[str] = set()
    silver_fund_isins: set[str] = set()
    for i in held_funds:
        name_to_check = " ".join(filter(None, [i.tradingsymbol, i.name]))
        if i.isin and _DEBT_KEYWORDS.search(name_to_check):
            debt_fund_isins.add(i.isin)
        if i.isin and _ARBITRAGE_RE.search(name_to_check):
            arbitrage_fund_isins.add(i.isin)
        if i.isin and _GOLD_RE.search(name_to_check):
            gold_fund_isins.add(i.isin)
        if i.isin and _SILVER_RE.search(name_to_check):
            silver_fund_isins.add(i.isin)

    if not BREAKDOWN_DIR.exists():
        return {"schemes_processed": 0, "rows_upserted": 0, "unmatched_equities": [],
                "missing_funds": [], "errors": ["Directory data/mf_portfolio_breakdown/ not found"]}

    csv_files = [p for p in BREAKDOWN_DIR.glob("*.csv") if p.stem.strip().startswith("IN")]
    schemes_processed = 0
    rows_upserted = 0
    unmatched: list[dict] = []
    errors: list[str] = []
    seen_isins: set[str] = set()

    isin_to_name = {i.isin: (i.tradingsymbol or i.name or i.isin) for i in held_funds if i.isin}
    if on_progress:
        await on_progress(f"Found {len(csv_files)} CSV file(s) starting with IN")

    for csv_path in csv_files:
        scheme_isin = csv_path.stem.strip()
        fund_name = isin_to_name.get(scheme_isin, scheme_isin)
        if on_progress:
            await on_progress(f"[{schemes_processed + 1}/{len(csv_files)}] {fund_name}")

        seen_isins.add(scheme_isin)
        is_debt_fund = scheme_isin in debt_fund_isins
        is_gold_fund = scheme_isin in gold_fund_isins
        is_silver_fund = scheme_isin in silver_fund_isins
        is_foreign_fund = scheme_isin in FOREIGN_FUND_ISINS
        equity_override = ETF_CAP_OVERRIDE.get(scheme_isin)

        try:
            text = csv_path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            text = csv_path.read_text(encoding="latin-1")

        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            errors.append(f"{csv_path.name}: empty or no header")
            continue
        reader.fieldnames = [f.strip() for f in reader.fieldnames]

        dedup: dict[tuple[str, str], dict] = {}
        for row_num, row in enumerate(reader, start=2):
            name = (row.get("Name") or "").strip()
            htype = (row.get("Type") or "").strip()
            pct_raw = row.get("Holdings") or ""
            if not name or not htype:
                continue
            pct = _parse_holdings_pct(pct_raw)
            if pct is None:
                errors.append(f"{csv_path.name} row {row_num}: bad Holdings '{pct_raw}'")
                continue

            if scheme_isin in arbitrage_fund_isins and htype.strip() in ("Equity - Future", "Cash - General Offset"):
                continue

            key = (name[:255], htype[:50])
            if key in dedup:
                dedup[key]["holdings_pct"] += pct
            else:
                is_arb_holding = scheme_isin in arbitrage_fund_isins and htype.strip() == "Equity"
                is_reit = bool(_REIT_RE.search(name) or _REIT_RE.search(htype))
                if is_reit:
                    category = "Real Estate Trust"
                elif is_gold_fund:
                    category = "Gold"
                elif is_silver_fund:
                    category = "Silver"
                elif is_debt_fund:
                    category = "Debt"
                elif is_arb_holding:
                    category = "Equity - Arbitrage"
                elif htype.strip() == "Equity":
                    name_lower = name.lower()
                    if is_foreign_fund or any(s in name_lower for s in FOREIGN_COMPANY_SUBSTRINGS):
                        category = "Equity - Foreign"
                    else:
                        category = _resolve_equity_category(name, alias_to_isin, name_to_isin, isin_to_mcap, amfi_by_name)
                        if category == "Unclassified Equity":
                            # AMFI has no record — fall back to the fund's index cap tier, then manual overrides
                            category = overrides.get(normalize_company_name(name)) or equity_override or "Unclassified Equity"
                else:
                    category = _classify_type(htype, name)
                if category == "Unclassified Equity":
                    unmatched.append({"name": name, "scheme_isin": scheme_isin})

                # Sector classification
                if is_reit:
                    sector: str | None = "Real Estate Trust"
                elif is_gold_fund:
                    sector = "Gold"
                elif is_silver_fund:
                    sector = "Silver"
                else:
                    sector = _sector_for_type(htype, name)
                    if sector is None and htype.strip() == "Equity":
                        sector = _resolve_equity_sector(name, alias_to_isin, name_to_isin, isin_to_sector, amfi_name_sector)

                dedup[key] = {
                    "scheme_isin": scheme_isin,
                    "name": key[0],
                    "holding_type": key[1],
                    "holdings_pct": pct,
                    "category": category,
                    "sector": sector,
                    "updated_at": now_ist(),
                }

        values = list(dedup.values())
        await db.execute(delete(MfSchemeBreakdown).where(MfSchemeBreakdown.scheme_isin == scheme_isin))
        if values:
            await db.execute(MfSchemeBreakdown.__table__.insert(), values)
            rows_upserted += len(values)

        schemes_processed += 1
        if on_progress:
            unmatched_here = sum(1 for u in unmatched if u["scheme_isin"] == scheme_isin)
            msg = f"  → {len(values)} rows"
            if unmatched_here:
                msg += f", {unmatched_here} unmatched"
            await on_progress(msg)

    # Synthesize 100% rows for commodity ETFs with no CSV (gold/silver trackers need no breakdown).
    for isin, commodity_cat in COMMODITY_ETF_CATEGORY.items():
        if isin not in held_isins or isin in seen_isins:
            continue
        await db.execute(delete(MfSchemeBreakdown).where(MfSchemeBreakdown.scheme_isin == isin))
        await db.execute(MfSchemeBreakdown.__table__.insert(), [{
            "scheme_isin": isin,
            "name": commodity_cat,
            "holding_type": commodity_cat,
            "holdings_pct": 100.0,
            "category": commodity_cat,
            "sector": commodity_cat,
            "updated_at": now_ist(),
        }])
        seen_isins.add(isin)
        rows_upserted += 1
        if on_progress:
            await on_progress(f"  → {isin_to_name.get(isin, isin)}: synthesized 100% {commodity_cat} (no CSV needed)")

    # Remove rows for schemes whose CSV was deleted.
    if seen_isins:
        await db.execute(
            delete(MfSchemeBreakdown).where(MfSchemeBreakdown.scheme_isin.notin_(seen_isins))
        )
    else:
        await db.execute(delete(MfSchemeBreakdown))

    await db.commit()

    # Warn about held funds that have no CSV file.
    missing_funds = [
        {"isin": isin, "name": isin_to_name.get(isin, isin)}
        for isin in sorted(held_isins - seen_isins)
    ]

    return {
        "schemes_processed": schemes_processed,
        "rows_upserted": rows_upserted,
        "unmatched_equities": unmatched,
        "missing_funds": missing_funds,
        "errors": errors[:30],
    }


_SGB_RE = re.compile(r"^SGB", re.IGNORECASE)


async def _load_amfi_lookups(db: AsyncSession) -> tuple[dict[str, str], dict[str, str]]:
    """Returns (isin_to_cat, norm_name_to_cat) from AmfiMarketCap."""
    amfi_rows = (await db.execute(select(AmfiMarketCap))).scalars().all()
    isin_to_cat: dict[str, str] = {}
    name_to_cat: dict[str, str] = {}
    for a in amfi_rows:
        if a.isin:
            isin_to_cat[a.isin] = a.categorization
        name_to_cat[a.name_normalized] = a.categorization
    return isin_to_cat, name_to_cat


def _classify_stock_instrument(
    isin: str | None,
    name: str | None,
    tradingsymbol: str | None,
    isin_to_cat: dict[str, str],
    name_to_cat: dict[str, str],
) -> str:
    if isin and isin in isin_to_cat:
        return isin_to_cat[isin]
    for raw in (name, tradingsymbol):
        if not raw:
            continue
        norm = normalize_company_name(raw)
        cat = name_to_cat.get(norm)
        if cat:
            return cat
        best_ratio = 0.0
        best_cat = None
        for amfi_norm, amfi_cat in name_to_cat.items():
            r = SequenceMatcher(None, norm, amfi_norm).ratio()
            if r > best_ratio:
                best_ratio = r
                best_cat = amfi_cat
        if best_ratio >= 0.85 and best_cat:
            return best_cat
    return "Unclassified Equity"


async def get_stock_holdings_table(db: AsyncSession) -> list[dict]:
    isin_to_cat, name_to_cat = await _load_amfi_lookups(db)

    # Build ticker lookup from AMFI
    amfi_rows = (await db.execute(select(AmfiMarketCap))).scalars().all()
    ticker_lookup: dict[str, str] = {}
    for a in amfi_rows:
        ticker_lookup[normalize_company_name(a.company_name)] = a.nse_symbol or a.bse_symbol or ""

    # MF/ETF fund holdings
    fund_result = await db.execute(
        select(Holding, Instrument)
        .join(Instrument, Holding.instrument_id == Instrument.id)
        .where(Instrument.instrument_type.in_(("MF", "ETF")))
    )
    fund_holdings = fund_result.all()

    fund_values: dict[str, float] = {}
    for h, i in fund_holdings:
        if not i.isin:
            continue
        ltp = float(h.last_price) if h.last_price else None
        fund_values[i.isin] = float(h.quantity) * ltp if ltp else float(h.total_cost or 0)

    stock_totals: dict[str, dict] = {}

    if fund_values:
        breakdown_rows = (await db.execute(
            select(MfSchemeBreakdown).where(
                MfSchemeBreakdown.scheme_isin.in_(list(fund_values.keys())),
                MfSchemeBreakdown.holding_type == "Equity",
            )
        )).scalars().all()

        for row in breakdown_rows:
            hv = fund_values.get(row.scheme_isin, 0)
            contribution = hv * (float(row.holdings_pct) / 100.0)
            if row.name not in stock_totals:
                ticker = ticker_lookup.get(normalize_company_name(row.name), "")
                stock_totals[row.name] = {"ticker": ticker, "category": row.category, "value": 0}
            stock_totals[row.name]["value"] += contribution

    # Direct stock holdings
    stock_result = await db.execute(
        select(Holding, Instrument)
        .join(Instrument, Holding.instrument_id == Instrument.id)
        .where(Instrument.instrument_type == "STOCK")
    )
    for h, i in stock_result.all():
        ltp = float(h.last_price) if h.last_price else None
        value = float(h.quantity) * ltp if ltp else float(h.total_cost or 0)
        name = i.name or i.tradingsymbol or "Unknown"
        cat = _classify_stock_instrument(i.isin, i.name, i.tradingsymbol, isin_to_cat, name_to_cat)
        ticker = i.tradingsymbol or ""
        if name in stock_totals:
            stock_totals[name]["value"] += value
        else:
            stock_totals[name] = {"ticker": ticker, "category": cat, "value": value}

    total_equity = sum(s["value"] for s in stock_totals.values())
    if total_equity <= 0:
        return []

    stocks = []
    for name, info in stock_totals.items():
        if info["value"] <= 0:
            continue
        stocks.append({
            "name": name,
            "ticker": info["ticker"],
            "category": info["category"],
            "weight_pct": round(info["value"] / total_equity * 100, 4),
            "value": round(info["value"], 2),
        })

    stocks.sort(key=lambda s: s["value"], reverse=True)
    return stocks


async def _build_category_totals_full(db: AsyncSession, all_holdings, use_cost: bool) -> dict[str, float]:
    from app.services.manual_assets import get_manual_assets_summary

    isin_to_cat, name_to_cat = await _load_amfi_lookups(db)
    category_totals: dict[str, float] = {}
    fund_isins: list[str] = []

    for h, i in all_holdings:
        if use_cost:
            value = float(h.total_cost or 0)
        else:
            ltp = float(h.last_price) if h.last_price else None
            value = float(h.quantity) * ltp if ltp else float(h.total_cost or 0)

        if i.instrument_type == "STOCK":
            cat = _classify_stock_instrument(i.isin, i.name, i.tradingsymbol, isin_to_cat, name_to_cat)
            category_totals[cat] = category_totals.get(cat, 0) + value
        elif i.instrument_type == "BOND" and i.tradingsymbol and _SGB_RE.match(i.tradingsymbol):
            category_totals["Gold"] = category_totals.get("Gold", 0) + value
        elif i.instrument_type == "BOND":
            category_totals["Debt"] = category_totals.get("Debt", 0) + value
        elif i.instrument_type in ("MF", "ETF") and i.isin:
            commodity_cat = COMMODITY_ETF_CATEGORY.get(i.isin)
            if commodity_cat:
                category_totals[commodity_cat] = category_totals.get(commodity_cat, 0) + value
            else:
                fund_isins.append(i.isin)

    hv: dict[str, float] = {}
    for h, i in all_holdings:
        if i.isin and i.isin in fund_isins:
            if use_cost:
                hv[i.isin] = float(h.total_cost or 0)
            else:
                ltp = float(h.last_price) if h.last_price else None
                hv[i.isin] = float(h.quantity) * ltp if ltp else float(h.total_cost or 0)

    if hv:
        breakdown_rows = (await db.execute(
            select(MfSchemeBreakdown).where(
                MfSchemeBreakdown.scheme_isin.in_(list(hv.keys()))
            )
        )).scalars().all()
        for row in breakdown_rows:
            contribution = hv.get(row.scheme_isin, 0) * (float(row.holdings_pct) / 100.0)
            category_totals[row.category] = category_totals.get(row.category, 0) + contribution

    manual = await get_manual_assets_summary(db)
    debt = manual["total_fd"] + manual["total_ppf"]
    if manual["nps"]:
        nps_val = manual["nps"]["current_value"]
        category_totals["Large Cap"] = category_totals.get("Large Cap", 0) + nps_val * 0.75
        debt += nps_val * 0.25
    if debt > 0:
        category_totals["Debt"] = category_totals.get("Debt", 0) + debt
    if manual.get("total_cash", 0) > 0:
        category_totals["Cash"] = category_totals.get("Cash", 0) + manual["total_cash"]
    if manual.get("total_foreign_equity_inr", 0) > 0:
        category_totals["Equity - Foreign"] = category_totals.get("Equity - Foreign", 0) + manual["total_foreign_equity_inr"]

    return category_totals


async def get_breakdown_chart_data(db: AsyncSession) -> dict:
    result = await db.execute(
        select(Holding, Instrument)
        .join(Instrument, Holding.instrument_id == Instrument.id)
        .where(Instrument.instrument_type.in_(("MF", "ETF", "BOND", "STOCK")))
    )
    all_holdings = result.all()
    category_totals = await _build_category_totals_full(db, all_holdings, use_cost=False)

    if not category_totals:
        return {"labels": [], "values": [], "total": 0}

    order = [
        "Large Cap", "Mid Cap", "Small Cap", "Unclassified Equity",
        "Equity - Foreign", "Equity - Arbitrage", "Real Estate Trust", "Gold", "Silver", "Debt", "Cash", "Other",
    ]
    labels = []
    values = []
    for cat in order:
        v = category_totals.get(cat, 0)
        if v > 0:
            labels.append(cat)
            values.append(round(v, 2))

    return {
        "labels": labels,
        "values": values,
        "total": round(sum(values), 2),
    }


DEFAULT_TARGETS = {
    "Large Cap": 50.0,
    "Mid Cap": 30.0,
    "Small Cap": 20.0,
    "Equity - Foreign": 0.0,
}


async def get_allocation_targets(db: AsyncSession) -> dict[str, float]:
    rows = (await db.execute(select(AllocationTarget))).scalars().all()
    if not rows:
        return dict(DEFAULT_TARGETS)
    return {r.category: float(r.target_pct) for r in rows}


async def save_allocation_targets(db: AsyncSession, targets: dict[str, float]):
    for category, pct in targets.items():
        existing = (await db.execute(
            select(AllocationTarget).where(AllocationTarget.category == category)
        )).scalar_one_or_none()
        if existing:
            existing.target_pct = pct
        else:
            db.add(AllocationTarget(category=category, target_pct=pct))
    await db.execute(
        delete(AllocationTarget).where(
            AllocationTarget.category.notin_(list(targets.keys()))
        )
    )
    await db.commit()


DEFAULT_ASSET_CLASS_TARGETS: dict[str, float] = {
    "Equity": 65.0,
    "Debt": 30.0,
    "Precious Metals": 5.0,
}


async def get_asset_class_targets(db: AsyncSession) -> dict[str, float]:
    rows = (await db.execute(select(AssetClassTarget))).scalars().all()
    if not rows:
        return dict(DEFAULT_ASSET_CLASS_TARGETS)
    return {r.asset_class: float(r.target_pct) for r in rows}


async def save_asset_class_targets(db: AsyncSession, targets: dict[str, float]):
    for asset_class, pct in targets.items():
        existing = (await db.execute(
            select(AssetClassTarget).where(AssetClassTarget.asset_class == asset_class)
        )).scalar_one_or_none()
        if existing:
            existing.target_pct = pct
        else:
            db.add(AssetClassTarget(asset_class=asset_class, target_pct=pct))
    await db.execute(
        delete(AssetClassTarget).where(
            AssetClassTarget.asset_class.notin_(list(targets.keys()))
        )
    )
    await db.commit()


_AC_EQUITY = {"Large Cap", "Mid Cap", "Small Cap", "Unclassified Equity", "Equity - Foreign"}
_AC_DEBT = {"Debt", "Equity - Arbitrage"}
_AC_PRECIOUS_METALS = {"Gold", "Silver"}


async def get_asset_class_comparison(db: AsyncSession) -> dict:
    from app.services.manual_assets import get_manual_assets_summary

    result = await db.execute(
        select(Holding, Instrument)
        .join(Instrument, Holding.instrument_id == Instrument.id)
        .where(Instrument.instrument_type.in_(("MF", "ETF", "BOND", "STOCK")))
    )
    all_holdings = result.all()
    category_totals = await _build_category_totals_full(db, all_holdings, use_cost=False)
    manual = await get_manual_assets_summary(db)

    savings_cash = manual.get("total_cash", 0)
    emergency_fund = manual.get("emergency_total", 0)
    ppf = manual.get("total_ppf", 0)

    # MF internal cash = total "Cash" category minus savings-account-only cash
    mf_cash = max(0.0, category_totals.get("Cash", 0) - savings_cash)

    equity = sum(category_totals.get(c, 0) for c in _AC_EQUITY)
    debt = sum(category_totals.get(c, 0) for c in _AC_DEBT) + mf_cash - emergency_fund - ppf
    debt = max(0.0, debt)
    precious_metals = sum(category_totals.get(c, 0) for c in _AC_PRECIOUS_METALS)

    investable_total = equity + debt + precious_metals
    grand_total = investable_total + savings_cash + emergency_fund + ppf

    targets = await get_asset_class_targets(db)

    foreign_target_row = (await db.execute(
        select(AllocationTarget).where(AllocationTarget.category == "Equity - Foreign")
    )).scalar_one_or_none()
    foreign_equity_target = float(foreign_target_row.target_pct) if foreign_target_row else 20.0

    rows = []
    for asset_class, current_value in [
        ("Equity", equity),
        ("Debt", debt),
        ("Precious Metals", precious_metals),
    ]:
        target_pct = targets.get(asset_class, DEFAULT_ASSET_CLASS_TARGETS.get(asset_class, 0.0))
        current_pct = (current_value / investable_total * 100) if investable_total > 0 else 0.0
        current_diff = current_pct - target_pct
        ideal_value = investable_total * target_pct / 100 if investable_total > 0 else 0.0
        shortfall = current_value - ideal_value
        rows.append({
            "asset_class": asset_class,
            "target_pct": target_pct,
            "current_pct": round(current_pct, 2),
            "current_value": round(current_value, 2),
            "current_diff": round(current_diff, 2),
            "ideal_value": round(ideal_value, 2),
            "shortfall": round(shortfall, 2),
        })

    return {
        "rows": rows,
        "foreign_equity_target": foreign_equity_target,
        "investable_total": round(investable_total, 2),
        "excluded": {
            "emergency_fund": round(emergency_fund, 2),
            "ppf": round(ppf, 2),
            "cash": round(savings_cash, 2),
            "total_excluded": round(savings_cash + emergency_fund + ppf, 2),
        },
        "grand_total": round(grand_total, 2),
    }


DOMESTIC_EQUITY_CATS = {"Large Cap", "Mid Cap", "Small Cap", "Unclassified Equity"}
FOREIGN_CAT = "Equity - Foreign"
EQUITY_CATS = DOMESTIC_EQUITY_CATS | {FOREIGN_CAT}


def _foreign_anchor_ratio(foreign_target: float, large_target: float) -> float:
    """Ratio of foreign ideal value to large cap value in anchored mode.
    Derived from: foreign_frac / (large_frac * domestic_frac)
    e.g. with foreign=20%, large=50%: 0.20 / (0.50 * 0.80) = 0.50
    """
    foreign_frac = foreign_target / 100
    large_frac = large_target / 100
    domestic_frac = 1.0 - foreign_frac
    if large_frac == 0 or domestic_frac == 0:
        return 0.5
    return foreign_frac / (large_frac * domestic_frac)


async def get_allocation_comparison(db: AsyncSession, mode: str = "anchored") -> dict:
    targets = await get_allocation_targets(db)

    result = await db.execute(
        select(Holding, Instrument)
        .join(Instrument, Holding.instrument_id == Instrument.id)
        .where(Instrument.instrument_type.in_(("MF", "ETF", "BOND", "STOCK")))
    )
    all_holdings = result.all()

    current_totals = await _build_category_totals_full(db, all_holdings, use_cost=False)
    invested_totals = await _build_category_totals_full(db, all_holdings, use_cost=True)

    current_equity = sum(v for c, v in current_totals.items() if c in EQUITY_CATS)
    invested_equity = sum(v for c, v in invested_totals.items() if c in EQUITY_CATS)

    foreign_cur = current_totals.get(FOREIGN_CAT, 0)
    foreign_inv = invested_totals.get(FOREIGN_CAT, 0)
    domestic_cur = current_equity - foreign_cur
    domestic_inv = invested_equity - foreign_inv

    foreign_target = targets.get(FOREIGN_CAT, 0)
    domestic_target = 100.0 - foreign_target

    large_target = targets.get("Large Cap", 50)
    cur_large = current_totals.get("Large Cap", 0)
    inv_large = invested_totals.get("Large Cap", 0)

    # ── Domestic market-cap rows ──────────────────────────────────────────────
    domestic_categories = sorted(c for c in targets if c in DOMESTIC_EQUITY_CATS)
    rows = []
    for cat in domestic_categories:
        target_pct = targets[cat]
        cur_val = current_totals.get(cat, 0)
        inv_val = invested_totals.get(cat, 0)

        if mode == "anchored":
            # Percentages relative to domestic equity; ideal anchored on large cap
            cur_pct = (cur_val / domestic_cur * 100) if domestic_cur > 0 else 0
            inv_pct = (inv_val / domestic_inv * 100) if domestic_inv > 0 else 0
            if cat == "Large Cap":
                cur_ideal_val = cur_large
                inv_ideal_val = inv_large
            else:
                cur_ideal_val = cur_large * (target_pct / large_target) if large_target > 0 else 0
                inv_ideal_val = inv_large * (target_pct / large_target) if large_target > 0 else 0
        else:  # free_float
            # Percentages relative to total equity; ideal = total_equity × target_of_total
            domestic_share = 1.0 - foreign_target / 100
            target_of_total = target_pct * domestic_share
            cur_pct = (cur_val / current_equity * 100) if current_equity > 0 else 0
            inv_pct = (inv_val / invested_equity * 100) if invested_equity > 0 else 0
            cur_ideal_val = current_equity * target_of_total / 100
            inv_ideal_val = invested_equity * target_of_total / 100
            target_pct = round(target_of_total, 2)  # expose as % of total equity in this mode

        rows.append({
            "category": cat,
            "target_pct": target_pct,
            "current_pct": round(cur_pct, 2),
            "current_value": round(cur_val, 2),
            "current_diff": round(cur_pct - target_pct, 2),
            "invested_pct": round(inv_pct, 2),
            "invested_value": round(inv_val, 2),
            "invested_diff": round(inv_pct - target_pct, 2),
            "current_ideal_value": round(cur_ideal_val, 2),
            "current_value_diff": round(cur_val - cur_ideal_val, 2),
            "invested_ideal_value": round(inv_ideal_val, 2),
            "invested_value_diff": round(inv_val - inv_ideal_val, 2),
        })

    # ── Foreign row (merged into rows) ───────────────────────────────────────
    foreign_cur_pct = (foreign_cur / current_equity * 100) if current_equity > 0 else 0
    foreign_inv_pct = (foreign_inv / invested_equity * 100) if invested_equity > 0 else 0

    if mode == "anchored":
        anchor_ratio = _foreign_anchor_ratio(foreign_target, large_target)
        cur_foreign_ideal = cur_large * anchor_ratio
        inv_foreign_ideal = inv_large * anchor_ratio
        # target_pct is the equivalent % of total equity for display
        foreign_display_target = round(cur_foreign_ideal / current_equity * 100, 2) if current_equity > 0 else 0
    else:
        cur_foreign_ideal = current_equity * foreign_target / 100
        inv_foreign_ideal = invested_equity * foreign_target / 100
        foreign_display_target = foreign_target

    rows.append({
        "category": "Equity - Foreign",
        "target_pct": foreign_display_target,
        "anchor_note": f"{anchor_ratio * 100:.1f}% of LC" if mode == "anchored" else None,
        "current_pct": round(foreign_cur_pct, 2),
        "current_value": round(foreign_cur, 2),
        "current_diff": round(foreign_cur_pct - foreign_display_target, 2),
        "invested_pct": round(foreign_inv_pct, 2),
        "invested_value": round(foreign_inv, 2),
        "invested_diff": round(foreign_inv_pct - foreign_display_target, 2),
        "current_ideal_value": round(cur_foreign_ideal, 2),
        "current_value_diff": round(foreign_cur - cur_foreign_ideal, 2),
        "invested_ideal_value": round(inv_foreign_ideal, 2),
        "invested_value_diff": round(foreign_inv - inv_foreign_ideal, 2),
    })

    # ── Domestic / foreign split summaries (kept for backward compat) ─────────
    domestic_cur_pct = (domestic_cur / current_equity * 100) if current_equity > 0 else 0
    domestic_inv_pct = (domestic_inv / invested_equity * 100) if invested_equity > 0 else 0
    domestic_cur_ideal = current_equity * domestic_target / 100
    foreign = {
        "target_pct": foreign_target,
        "current_pct": round(foreign_cur_pct, 2),
        "current_value": round(foreign_cur, 2),
        "current_diff": round(foreign_cur_pct - foreign_target, 2),
        "current_value_diff": round(foreign_cur - cur_foreign_ideal, 2),
        "invested_pct": round(foreign_inv_pct, 2),
        "invested_value": round(foreign_inv, 2),
    }
    domestic = {
        "target_pct": domestic_target,
        "current_pct": round(domestic_cur_pct, 2),
        "current_value": round(domestic_cur, 2),
        "current_diff": round(domestic_cur_pct - domestic_target, 2),
        "current_value_diff": round(domestic_cur - domestic_cur_ideal, 2),
        "invested_pct": round(domestic_inv_pct, 2),
        "invested_value": round(domestic_inv, 2),
    }

    return {
        "rows": rows,
        "foreign": foreign,
        "domestic": domestic,
        "targets": targets,
        "current_equity": round(current_equity, 2),
        "invested_equity": round(invested_equity, 2),
        "domestic_equity": round(domestic_cur, 2),
        "mode": mode,
    }


async def get_available_schemes(db: AsyncSession) -> list[dict]:
    scheme_isins = (await db.execute(
        select(MfSchemeBreakdown.scheme_isin).distinct()
    )).scalars().all()
    if not scheme_isins:
        return []

    instruments = (await db.execute(
        select(Instrument).where(Instrument.isin.in_(scheme_isins))
    )).scalars().all()
    isin_to_name = {i.isin: i.tradingsymbol or i.name or i.isin for i in instruments}

    result = []
    for isin in sorted(scheme_isins, key=lambda s: isin_to_name.get(s, s)):
        result.append({"scheme_isin": isin, "name": isin_to_name.get(isin, isin)})
    return result


async def get_scheme_breakdown(db: AsyncSession, scheme_isin: str) -> dict:
    rows = (await db.execute(
        select(MfSchemeBreakdown)
        .where(MfSchemeBreakdown.scheme_isin == scheme_isin)
        .order_by(MfSchemeBreakdown.holdings_pct.desc())
    )).scalars().all()

    if not rows:
        return {"holdings": [], "category_summary": []}

    # Resolve fund market value from the holding record
    holding_row = (await db.execute(
        select(Holding, Instrument)
        .join(Instrument, Holding.instrument_id == Instrument.id)
        .where(Instrument.isin == scheme_isin)
    )).first()
    fund_value = 0.0
    if holding_row:
        h, _ = holding_row
        ltp = float(h.last_price) if h.last_price else None
        fund_value = float(h.quantity) * ltp if ltp else float(h.total_cost or 0)

    holdings = []
    cat_value_totals: dict[str, float] = {}
    cat_pct_totals: dict[str, float] = {}
    for r in rows:
        pct = float(r.holdings_pct)
        value = round(fund_value * (pct / 100.0), 2)
        holdings.append({
            "name": r.name,
            "type": r.holding_type,
            "category": r.category,
            "pct": round(pct, 4),
            "value": value,
        })
        cat_value_totals[r.category] = cat_value_totals.get(r.category, 0) + value
        cat_pct_totals[r.category] = cat_pct_totals.get(r.category, 0) + pct

    order = [
        "Large Cap", "Mid Cap", "Small Cap", "Unclassified Equity",
        "Equity - Foreign", "Equity - Arbitrage", "Real Estate Trust", "Gold", "Silver", "Debt", "Cash", "Other",
    ]
    category_summary = []
    for cat in order:
        pct_total = cat_pct_totals.get(cat, 0)
        if pct_total > 0:
            category_summary.append({
                "category": cat,
                "pct": round(pct_total, 2),
                "value": round(cat_value_totals.get(cat, 0), 2),
            })

    return {"holdings": holdings, "category_summary": category_summary}


_CAT_ORDER = ["Large Cap", "Mid Cap", "Small Cap", "Unclassified Equity", "Equity - Foreign", "Equity - Arbitrage", "Real Estate Trust", "Gold", "Silver", "Debt", "Cash", "Other"]


async def get_category_composition(db: AsyncSession) -> list[dict]:
    """Returns per-category breakdown showing each contributing source and its value."""
    from app.services.manual_assets import get_manual_assets_summary

    result = await db.execute(
        select(Holding, Instrument)
        .join(Instrument, Holding.instrument_id == Instrument.id)
        .where(Instrument.instrument_type.in_(("MF", "ETF", "BOND", "STOCK")))
    )
    all_holdings = result.all()

    isin_to_cat, name_to_cat = await _load_amfi_lookups(db)

    # cat -> list of {name, source_type, fund_pct, contribution}
    composition: dict[str, list[dict]] = {}

    def _add(cat: str, entry: dict):
        composition.setdefault(cat, []).append(entry)

    # MF/ETF: group breakdown rows by (scheme_isin, category)
    fund_values: dict[str, tuple[float, str]] = {}
    for h, i in all_holdings:
        if i.instrument_type in ("MF", "ETF") and i.isin:
            ltp = float(h.last_price) if h.last_price else None
            val = float(h.quantity) * ltp if ltp else float(h.total_cost or 0)
            commodity_cat = COMMODITY_ETF_CATEGORY.get(i.isin)
            if commodity_cat:
                _add(commodity_cat, {"name": i.name or i.tradingsymbol or i.isin, "source_type": "etf", "fund_pct": 100.0, "contribution": round(val, 2)})
            else:
                fund_values[i.isin] = (val, i.name or i.tradingsymbol or i.isin)

    if fund_values:
        breakdown_rows = (await db.execute(
            select(MfSchemeBreakdown).where(MfSchemeBreakdown.scheme_isin.in_(list(fund_values.keys())))
        )).scalars().all()

        scheme_cat: dict[tuple, float] = defaultdict(float)
        for row in breakdown_rows:
            scheme_cat[(row.scheme_isin, row.category)] += float(row.holdings_pct)

        for (isin, cat), pct in scheme_cat.items():
            fund_val, fund_name = fund_values[isin]
            contribution = fund_val * pct / 100
            if contribution <= 0:
                continue
            _add(cat, {"name": fund_name, "isin": isin, "source_type": "fund", "fund_pct": round(pct, 2), "contribution": round(contribution, 2), "fund_value": round(fund_val, 2)})

    # Direct stocks
    for h, i in all_holdings:
        if i.instrument_type == "STOCK":
            ltp = float(h.last_price) if h.last_price else None
            val = float(h.quantity) * ltp if ltp else float(h.total_cost or 0)
            if val <= 0:
                continue
            cat = _classify_stock_instrument(i.isin, i.name, i.tradingsymbol, isin_to_cat, name_to_cat)
            _add(cat, {"name": i.name or i.tradingsymbol or "Unknown", "source_type": "stock", "fund_pct": 100.0, "contribution": round(val, 2)})

    # Bonds: SGB → Gold, everything else → Debt
    for h, i in all_holdings:
        if i.instrument_type == "BOND":
            ltp = float(h.last_price) if h.last_price else None
            val = float(h.quantity) * ltp if ltp else float(h.total_cost or 0)
            if val <= 0:
                continue
            if i.tradingsymbol and _SGB_RE.match(i.tradingsymbol):
                _add("Gold", {"name": i.tradingsymbol, "source_type": "bond", "fund_pct": 100.0, "contribution": round(val, 2)})
            else:
                _add("Debt", {"name": i.tradingsymbol or i.name or "Govt Bond", "source_type": "bond", "fund_pct": 100.0, "contribution": round(val, 2)})

    # Manual assets
    manual = await get_manual_assets_summary(db)
    if manual["total_fd"] > 0:
        _add("Debt", {"name": "Fixed Deposits", "source_type": "manual", "fund_pct": 100.0, "contribution": round(manual["total_fd"], 2)})
    if manual["total_ppf"] > 0:
        _add("Debt", {"name": "PPF", "source_type": "manual", "fund_pct": 100.0, "contribution": round(manual["total_ppf"], 2)})
    if manual.get("nps"):
        nps_val = manual["nps"]["current_value"]
        if nps_val > 0:
            _add("Large Cap", {"name": "NPS (equity portion)", "source_type": "manual", "fund_pct": 75.0, "contribution": round(nps_val * 0.75, 2)})
            _add("Debt", {"name": "NPS (debt portion)", "source_type": "manual", "fund_pct": 25.0, "contribution": round(nps_val * 0.25, 2)})
    if manual.get("total_cash", 0) > 0:
        _add("Cash", {"name": "Savings / Cash", "source_type": "manual", "fund_pct": 100.0, "contribution": round(manual["total_cash"], 2)})
    for fe in manual.get("foreign_equities", []):
        if fe["value_inr"] > 0:
            _add("Equity - Foreign", {"name": fe["label"], "source_type": "manual", "fund_pct": 100.0, "contribution": round(fe["value_inr"], 2)})

    # Build ordered list with totals and per-source share_pct
    out = []
    for cat in _CAT_ORDER:
        sources = composition.get(cat)
        if not sources:
            continue
        sources.sort(key=lambda x: x["contribution"], reverse=True)
        total = sum(s["contribution"] for s in sources)
        for s in sources:
            s["share_pct"] = round(s["contribution"] / total * 100, 1) if total else 0
        out.append({"category": cat, "total": round(total, 2), "sources": sources})

    return out


_NON_EQUITY_SECTORS = {"Fixed Income", "Liquid / Money Market", "Gold", "Silver"}


async def get_sector_composition(db: AsyncSession, equity_only: bool = False) -> list[dict]:
    """Returns per-sector breakdown showing each contributing source and its value."""
    from app.services.manual_assets import get_manual_assets_summary

    result = await db.execute(
        select(Holding, Instrument)
        .join(Instrument, Holding.instrument_id == Instrument.id)
        .where(Instrument.instrument_type.in_(("MF", "ETF", "BOND", "STOCK")))
    )
    all_holdings = result.all()

    # Sector lookups for direct stocks
    amfi_all = (await db.execute(select(AmfiMarketCap))).scalars().all()
    isin_to_sector: dict[str, str] = {a.isin: a.sector for a in amfi_all if a.isin and a.sector}
    name_to_sector: dict[str, str] = {a.name_normalized: a.sector for a in amfi_all if a.sector}

    composition: dict[str, list[dict]] = {}

    def _add(sector: str, entry: dict):
        composition.setdefault(sector, []).append(entry)

    # MF/ETF: group breakdown rows by (scheme_isin, sector)
    fund_values: dict[str, tuple[float, str]] = {}
    for h, i in all_holdings:
        if i.instrument_type in ("MF", "ETF") and i.isin:
            ltp = float(h.last_price) if h.last_price else None
            val = float(h.quantity) * ltp if ltp else float(h.total_cost or 0)
            commodity_cat = COMMODITY_ETF_CATEGORY.get(i.isin)
            if commodity_cat:
                _add(commodity_cat, {"name": i.name or i.tradingsymbol or i.isin, "source_type": "etf", "fund_pct": 100.0, "contribution": round(val, 2)})
            else:
                fund_values[i.isin] = (val, i.name or i.tradingsymbol or i.isin)

    if fund_values:
        breakdown_rows = (await db.execute(
            select(MfSchemeBreakdown).where(
                MfSchemeBreakdown.scheme_isin.in_(list(fund_values.keys())),
                MfSchemeBreakdown.category != "Equity - Arbitrage",
            )
        )).scalars().all()

        scheme_sector: dict[tuple, float] = defaultdict(float)
        for row in breakdown_rows:
            sec = row.sector or "Unknown"
            scheme_sector[(row.scheme_isin, sec)] += float(row.holdings_pct)

        for (isin, sec), pct in scheme_sector.items():
            fund_val, fund_name = fund_values[isin]
            contribution = fund_val * pct / 100
            if contribution <= 0:
                continue
            _add(sec, {"name": fund_name, "isin": isin, "source_type": "fund", "fund_pct": round(pct, 2), "contribution": round(contribution, 2), "fund_value": round(fund_val, 2)})

    # Direct stocks
    for h, i in all_holdings:
        if i.instrument_type == "STOCK":
            ltp = float(h.last_price) if h.last_price else None
            val = float(h.quantity) * ltp if ltp else float(h.total_cost or 0)
            if val <= 0:
                continue
            sec = isin_to_sector.get(i.isin or "")
            if sec is None and (i.name or i.tradingsymbol):
                norm = normalize_company_name(i.name or i.tradingsymbol or "")
                sec = name_to_sector.get(norm)
                if sec is None:
                    best_r, best_s = 0.0, None
                    for amfi_norm, amfi_sec in name_to_sector.items():
                        r = SequenceMatcher(None, norm, amfi_norm).ratio()
                        if r > best_r:
                            best_r, best_s = r, amfi_sec
                    if best_r >= 0.85:
                        sec = best_s
            _add(sec or "Unknown", {"name": i.name or i.tradingsymbol or "Unknown", "source_type": "stock", "fund_pct": 100.0, "contribution": round(val, 2)})

    if not equity_only:
        # Bonds: SGB → Gold sector, everything else → Fixed Income
        for h, i in all_holdings:
            if i.instrument_type == "BOND":
                ltp = float(h.last_price) if h.last_price else None
                val = float(h.quantity) * ltp if ltp else float(h.total_cost or 0)
                if val <= 0:
                    continue
                if i.tradingsymbol and _SGB_RE.match(i.tradingsymbol):
                    _add("Gold", {"name": i.tradingsymbol, "source_type": "bond", "fund_pct": 100.0, "contribution": round(val, 2)})
                else:
                    _add("Fixed Income", {"name": i.tradingsymbol or i.name or "Govt Bond", "source_type": "bond", "fund_pct": 100.0, "contribution": round(val, 2)})

        # Manual assets
        manual = await get_manual_assets_summary(db)
        if manual["total_fd"] > 0:
            _add("Fixed Income", {"name": "Fixed Deposits", "source_type": "manual", "fund_pct": 100.0, "contribution": round(manual["total_fd"], 2)})
        if manual["total_ppf"] > 0:
            _add("Fixed Income", {"name": "PPF", "source_type": "manual", "fund_pct": 100.0, "contribution": round(manual["total_ppf"], 2)})
        if manual.get("nps"):
            nps_val = manual["nps"]["current_value"]
            if nps_val > 0:
                _add("Fixed Income", {"name": "NPS (debt portion)", "source_type": "manual", "fund_pct": 25.0, "contribution": round(nps_val * 0.25, 2)})
        if manual.get("total_cash", 0) > 0:
            _add("Liquid / Money Market", {"name": "Savings / Cash", "source_type": "manual", "fund_pct": 100.0, "contribution": round(manual["total_cash"], 2)})

    # Sort by total descending, excluding non-equity sectors when equity_only
    out = []
    for sec, sources in sorted(composition.items(), key=lambda x: sum(s["contribution"] for s in x[1]), reverse=True):
        if equity_only and sec in _NON_EQUITY_SECTORS:
            continue
        sources.sort(key=lambda x: x["contribution"], reverse=True)
        total = sum(s["contribution"] for s in sources)
        for s in sources:
            s["share_pct"] = round(s["contribution"] / total * 100, 1) if total else 0
        out.append({"sector": sec, "total": round(total, 2), "sources": sources})

    return out


async def get_sector_stock_breakdown(db: AsyncSession) -> list[dict]:
    """Per-sector breakdown listing underlying stock holdings aggregated across all funds and direct positions."""
    fund_result = await db.execute(
        select(Holding, Instrument)
        .join(Instrument, Holding.instrument_id == Instrument.id)
        .where(Instrument.instrument_type.in_(("MF", "ETF")), Instrument.isin.isnot(None))
    )
    fund_values: dict[str, float] = {}
    for h, i in fund_result.all():
        ltp = float(h.last_price) if h.last_price else None
        val = float(h.quantity) * ltp if ltp else float(h.total_cost or 0)
        if val > 0:
            fund_values[i.isin] = val

    sector_stocks: dict[str, dict[str, float]] = {}

    if fund_values:
        rows = (await db.execute(
            select(MfSchemeBreakdown).where(
                MfSchemeBreakdown.scheme_isin.in_(list(fund_values.keys())),
                MfSchemeBreakdown.category != "Equity - Arbitrage",
                or_(
                    MfSchemeBreakdown.sector.is_(None),
                    ~MfSchemeBreakdown.sector.in_(list(_NON_EQUITY_SECTORS)),
                ),
            )
        )).scalars().all()
        for row in rows:
            contrib = fund_values[row.scheme_isin] * float(row.holdings_pct) / 100
            if contrib <= 0:
                continue
            sec = row.sector or "Unknown"
            bucket = sector_stocks.setdefault(sec, {})
            bucket[row.name] = bucket.get(row.name, 0) + contrib

    amfi_all = (await db.execute(select(AmfiMarketCap))).scalars().all()
    isin_to_sector: dict[str, str] = {a.isin: a.sector for a in amfi_all if a.isin and a.sector}
    name_to_sector: dict[str, str] = {a.name_normalized: a.sector for a in amfi_all if a.sector}

    stock_result = await db.execute(
        select(Holding, Instrument)
        .join(Instrument, Holding.instrument_id == Instrument.id)
        .where(Instrument.instrument_type == "STOCK")
    )
    for h, i in stock_result.all():
        ltp = float(h.last_price) if h.last_price else None
        val = float(h.quantity) * ltp if ltp else float(h.total_cost or 0)
        if val <= 0:
            continue
        sec = isin_to_sector.get(i.isin or "")
        if sec is None:
            norm = normalize_company_name(i.name or i.tradingsymbol or "")
            sec = name_to_sector.get(norm)
            if sec is None:
                best_r, best_s = 0.0, None
                for amfi_norm, amfi_sec in name_to_sector.items():
                    r = SequenceMatcher(None, norm, amfi_norm).ratio()
                    if r > best_r:
                        best_r, best_s = r, amfi_sec
                if best_r >= 0.85:
                    sec = best_s
        if sec in _NON_EQUITY_SECTORS:
            continue
        sec = sec or "Unknown"
        name = i.name or i.tradingsymbol or "Unknown"
        bucket = sector_stocks.setdefault(sec, {})
        bucket[name] = bucket.get(name, 0) + val

    out = []
    for sec, stocks in sorted(sector_stocks.items(), key=lambda x: sum(x[1].values()), reverse=True):
        total = sum(stocks.values())
        holdings = sorted(
            [{"name": n, "value": round(v, 2), "pct": round(v / total * 100, 1)} for n, v in stocks.items()],
            key=lambda x: x["value"], reverse=True,
        )
        out.append({"sector": sec, "total": round(total, 2), "holdings": holdings})
    return out


async def get_direct_trade_breakdown(db: AsyncSession) -> list[dict]:
    trades_result = await db.execute(
        select(Trade, Instrument)
        .join(Instrument, Trade.instrument_id == Instrument.id)
        .where(Trade.trade_type.in_(["BUY", "SELL"]))
        .order_by(Instrument.tradingsymbol, Trade.trade_date)
    )
    rows = trades_result.all()
    if not rows:
        return []

    instrument_info: dict[int, dict] = {}
    # key: (instrument_id, trade_date, trade_type) → {qty, cost}
    instrument_buckets: dict[tuple, dict] = defaultdict(lambda: {"qty": 0.0, "cost": 0.0})

    for trade, instrument in rows:
        iid = instrument.id
        if iid not in instrument_info:
            instrument_info[iid] = {
                "symbol": instrument.tradingsymbol or "",
                "name": instrument.name or instrument.tradingsymbol or "",
                "instrument_type": instrument.instrument_type,
            }
        qty = float(trade.quantity)
        key = (iid, trade.trade_date, trade.trade_type)
        instrument_buckets[key]["qty"] += qty
        instrument_buckets[key]["cost"] += qty * float(trade.price)

    instrument_ids = list(instrument_info.keys())

    # LTP from holdings
    ltp_map: dict[int, float] = {}
    holdings_result = await db.execute(
        select(Holding.instrument_id, Holding.last_price)
        .where(Holding.instrument_id.in_(instrument_ids))
    )
    for iid, lp in holdings_result.all():
        if lp:
            ltp_map[iid] = float(lp)

    # Fall back to latest price_history close
    missing = [iid for iid in instrument_ids if iid not in ltp_map]
    if missing:
        latest_sub = (
            select(PriceHistory.instrument_id, func.max(PriceHistory.price_date).label("max_date"))
            .where(PriceHistory.instrument_id.in_(missing))
            .group_by(PriceHistory.instrument_id)
            .subquery()
        )
        ph_result = await db.execute(
            select(PriceHistory.instrument_id, PriceHistory.close)
            .join(latest_sub, and_(
                PriceHistory.instrument_id == latest_sub.c.instrument_id,
                PriceHistory.price_date == latest_sub.c.max_date,
            ))
        )
        for iid, close in ph_result.all():
            if close:
                ltp_map[iid] = float(close)

    # Fall back to latest nav_history for MFs/ETFs still missing
    missing = [iid for iid in instrument_ids if iid not in ltp_map]
    if missing:
        latest_nav_sub = (
            select(NavHistory.instrument_id, func.max(NavHistory.nav_date).label("max_date"))
            .where(NavHistory.instrument_id.in_(missing))
            .group_by(NavHistory.instrument_id)
            .subquery()
        )
        nav_result = await db.execute(
            select(NavHistory.instrument_id, NavHistory.nav)
            .join(latest_nav_sub, and_(
                NavHistory.instrument_id == latest_nav_sub.c.instrument_id,
                NavHistory.nav_date == latest_nav_sub.c.max_date,
            ))
        )
        for iid, nav in nav_result.all():
            if nav:
                ltp_map[iid] = float(nav)

    # Group buckets back by instrument, sorted by (date, type)
    from collections import defaultdict as _dd
    iid_trades: dict[int, list] = _dd(list)
    for (iid, td, ttype), agg in instrument_buckets.items():
        iid_trades[iid].append((td, ttype, agg))

    result = []
    for iid in sorted(instrument_ids, key=lambda x: instrument_info[x]["symbol"]):
        info = instrument_info[iid]
        current_price = ltp_map.get(iid)
        trade_rows = []
        for td, ttype, agg in sorted(iid_trades[iid], key=lambda x: (x[0], x[1])):
            qty = round(agg["qty"], 6)
            price = round(agg["cost"] / qty, 4) if qty else 0.0
            amount = round(agg["cost"], 2)
            current_value = round(qty * current_price, 2) if current_price else None
            if current_price and price:
                raw_pct = (current_price - price) / price * 100
                pct_change = round(-raw_pct if ttype == "SELL" else raw_pct, 2)
            else:
                pct_change = None
            trade_rows.append({
                "date": str(td),
                "trade_type": ttype,
                "qty": qty,
                "price": price,
                "amount": amount,
                "current_value": current_value,
                "pct_change": pct_change,
            })
        result.append({
            "symbol": info["symbol"],
            "name": info["name"],
            "instrument_type": info["instrument_type"],
            "current_price": current_price,
            "trades": trade_rows,
        })
    return result
