import csv
import io
import re
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path

import httpx
import openpyxl
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.holding import Holding
from app.models.instrument import Instrument
from app.models.mf_breakdown import AmfiMarketCap, EquityCategoryOverride, EquitySectorOverride, MfSchemeBreakdown
from app.services.mfapi_nav import resolve_scheme_codes
from app.time_util import now_ist

BREAKDOWN_DIR = Path("data/mf_portfolio_breakdown")
OPENFIN_BASE = "https://openfin.pocketedge.in/api/v1"

_AMFI_DATE_RE = re.compile(
    r"AverageMarketCapitalization(\d{1,2})(\w{3})(\d{4})",
    re.IGNORECASE,
)
_MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

# ETFs whose entire equity portfolio is a single market-cap category.
# Fallback cap category for holdings in these ETFs that AMFI doesn't recognise.
# AMFI always takes priority; this only applies when AMFI returns Unclassified Equity.
ETF_CAP_OVERRIDE: dict[str, str] = {
    "INF204KB14I2": "Large Cap",   # NIFTYBEES
    "INF732E01045": "Large Cap",   # JUNIORBEES (Nifty Next 50)
    "INF769K01IC9": "Mid Cap",     # MIDCAPETF
    "INF0R8F01141": "Small Cap",   # SML100CASE
}

# Funds whose equity holdings are entirely foreign; bypasses AMFI market-cap lookup.
FOREIGN_FUND_ISINS: set[str] = {
    "INF247L01AP3",   # MON100 / Nasdaq 100
}

# Individual company names (normalised, lowercase substrings) that are always foreign equity,
# regardless of which fund holds them.
FOREIGN_COMPANY_SUBSTRINGS: set[str] = {
    "alphabet", "amazon", "apple", "meta platforms", "microsoft",
}

# Commodity ETFs whose entire value maps to a single non-equity category.
# These bypass MF breakdown CSV lookup entirely.
COMMODITY_ETF_CATEGORY: dict[str, str] = {
    "INF109KC1Y56": "Silver",   # SILVERIETF
    "INF204KB17I5": "Gold",     # GOLDBEES
    "INF0R8F01042": "Gold",     # GOLDCASE
}

# Matches Sovereign Gold Bond tradingsymbols (used across allocation and composition).
_SGB_RE = re.compile(r"^SGB", re.IGNORECASE)

_BRACKET_RE = re.compile(r"[\[(].*?[\])]")
_GLUED_DOT = re.compile(r"\.(?=[a-z])")
_ABBREV_MAP = [
    (re.compile(r"\bltd\.?(?=\W|$)", re.IGNORECASE), "limited"),
    (re.compile(r"\bcorpn?\.?(?=\W|$)", re.IGNORECASE), "corporation"),
    (re.compile(r"\bpvt\.?(?=\W|$)", re.IGNORECASE), "private"),
    (re.compile(r"\bco\.?(?=\W|$)", re.IGNORECASE), "company"),
]
_SUFFIX_RE = re.compile(
    r"\b(limited|limted|company|corporation|ordinary\s+shares|private)\b",
    re.IGNORECASE,
)
_TRAILING_JUNK = re.compile(r"[\s.*\-]+$")
_MULTI_SPACE = re.compile(r"\s+")

_NAME_ALIASES: dict[str, str] = {
    "m.r.f.": "mrf",
    "m r f": "mrf",
}


def normalize_company_name(name: str) -> str:
    s = name.lower().strip()
    s = _BRACKET_RE.sub("", s)
    s = _GLUED_DOT.sub(". ", s)
    for pat, expansion in _ABBREV_MAP:
        s = pat.sub(expansion, s)
    s = _SUFFIX_RE.sub("", s)
    s = s.replace("(", " ").replace(")", " ").replace(".", " ")
    s = _TRAILING_JUNK.sub("", s)
    s = _MULTI_SPACE.sub(" ", s).strip()
    for alias, canonical in _NAME_ALIASES.items():
        s = s.replace(alias, canonical)
    return s


def _find_amfi_xlsx() -> Path | None:
    if not BREAKDOWN_DIR.exists():
        return None
    candidates = list(BREAKDOWN_DIR.glob("AverageMarketCapitalization*.xlsx"))
    if not candidates:
        return None
    candidates.sort(key=lambda p: _parse_amfi_date(p.name) or date.min, reverse=True)
    return candidates[0]


def _parse_amfi_date(filename: str) -> date | None:
    m = _AMFI_DATE_RE.search(filename)
    if not m:
        return None
    day, mon_str, year = int(m.group(1)), m.group(2).lower(), int(m.group(3))
    month = _MONTH_MAP.get(mon_str[:3])
    if not month:
        return None
    return date(year, month, day)


def _load_sector_master() -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    """Returns ({isin: sector}, {normalized_name: sector}, {nse_symbol: sector})."""
    path = BREAKDOWN_DIR / "sector_master.csv"
    if not path.exists():
        return {}, {}, {}
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        text = path.read_text(encoding="latin-1")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return {}, {}, {}
    reader.fieldnames = [f.strip() for f in reader.fieldnames]
    isin_sector: dict[str, str] = {}
    name_sector: dict[str, str] = {}
    symbol_sector: dict[str, str] = {}
    for row in reader:
        isin = (row.get("ISIN Code") or "").strip()
        sector = (row.get("Industry") or "").strip()
        company = (row.get("Company Name") or "").strip()
        symbol = (row.get("Symbol") or "").strip()
        if not sector:
            continue
        if isin:
            isin_sector[isin] = sector
        if company:
            name_sector[normalize_company_name(company)] = sector
        if symbol:
            symbol_sector[symbol] = sector
    return isin_sector, name_sector, symbol_sector


def _write_company_master(rows: list[dict]) -> None:
    master_path = BREAKDOWN_DIR / "company_master.csv"
    BREAKDOWN_DIR.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "isin", "canonical_name", "primary_ticker", "nse_symbol", "bse_symbol",
        "msei_symbol", "exchanges", "mcap_category", "sector", "aliases",
    ]
    cat_order = {"Large Cap": 0, "Mid Cap": 1, "Small Cap": 2}
    sorted_rows = sorted(rows, key=lambda r: (cat_order.get(r["categorization"], 9), r["name_normalized"]))
    with master_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in sorted_rows:
            writer.writerow({
                "isin": r.get("isin") or "",
                "canonical_name": r.get("company_name") or "",
                "primary_ticker": r.get("primary_ticker") or "",
                "nse_symbol": r.get("nse_symbol") or "",
                "bse_symbol": r.get("bse_symbol") or "",
                "msei_symbol": r.get("msei_symbol") or "",
                "exchanges": r.get("exchanges") or "",
                "mcap_category": r.get("categorization") or "",
                "sector": r.get("sector") or "",
                "aliases": r.get("aliases") or "",
            })


async def sync_amfi_market_cap(db: AsyncSession, on_progress=None) -> dict:
    xlsx_path = _find_amfi_xlsx()
    if not xlsx_path:
        return {"error": "No AverageMarketCapitalization*.xlsx found in data/mf_portfolio_breakdown/"}

    file_date = _parse_amfi_date(xlsx_path.name)
    stale_warning = None
    if file_date:
        age_days = (date.today() - file_date).days
        if age_days > 180:
            stale_warning = f"Classification file is {age_days} days old ({file_date:%d %b %Y}). Consider updating from AMFI website."

    # Preserve user-edited aliases from existing company_master.csv
    master_path = BREAKDOWN_DIR / "company_master.csv"
    preserved_aliases: dict[str, str] = {}
    if master_path.exists():
        try:
            text = master_path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            text = master_path.read_text(encoding="latin-1")
        r = csv.DictReader(io.StringIO(text))
        for mrow in r:
            isin_key = (mrow.get("isin") or "").strip()
            alias_val = (mrow.get("aliases") or "").strip()
            if isin_key and alias_val:
                preserved_aliases[isin_key] = alias_val

    if on_progress:
        await on_progress(f"Loading AMFI data from {xlsx_path.name}…")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_to_insert: list[dict] = []
    counts = {"Large Cap": 0, "Mid Cap": 0, "Small Cap": 0}

    _DASH = {"-", "–", "—"}

    for row in ws.iter_rows(min_row=3, values_only=True):
        company = str(row[1] or "").strip()
        isin = str(row[2] or "").strip() or None
        bse_raw = str(row[3] or "").strip()
        nse_raw = str(row[5] or "").strip()
        msei_raw = str(row[7] or "").strip()
        cat = str(row[10] or "").strip()
        if not company or cat not in counts:
            continue

        bse_sym = bse_raw if bse_raw and bse_raw not in _DASH else None
        nse_sym = nse_raw if nse_raw and nse_raw not in _DASH else None
        msei_sym = msei_raw if msei_raw and msei_raw not in _DASH else None

        primary_ticker = nse_sym or bse_sym or msei_sym

        exchanges_parts: list[str] = []
        if nse_sym:
            exchanges_parts.append("NSE")
        if bse_sym:
            exchanges_parts.append("BSE")
        if msei_sym:
            exchanges_parts.append("MSEI")
        exchanges = ",".join(exchanges_parts) or None

        counts[cat] += 1
        rows_to_insert.append({
            "company_name": company,
            "isin": isin,
            "bse_symbol": bse_sym,
            "nse_symbol": nse_sym,
            "msei_symbol": msei_sym,
            "primary_ticker": primary_ticker,
            "exchanges": exchanges,
            "aliases": preserved_aliases.get(isin or "") or None,
            "categorization": cat,
            "sector": None,
            "name_normalized": normalize_company_name(company),
            "updated_at": now_ist(),
        })

    wb.close()

    if on_progress:
        await on_progress(
            f"Parsed {len(rows_to_insert)} companies "
            f"(Large: {counts['Large Cap']}, Mid: {counts['Mid Cap']}, Small: {counts['Small Cap']})"
        )

    # Enrich with sector: ISIN → NSE symbol → normalized name
    isin_sector, name_sector, symbol_sector = _load_sector_master()
    sectors_loaded = len(isin_sector)
    if isin_sector or name_sector or symbol_sector:
        for row in rows_to_insert:
            sec = isin_sector.get(row["isin"] or "")
            if not sec and row["nse_symbol"]:
                sec = symbol_sector.get(row["nse_symbol"])
            if not sec:
                sec = name_sector.get(row["name_normalized"] or "")
            if sec:
                row["sector"] = sec
        if on_progress:
            await on_progress(f"Sector enrichment: {sectors_loaded} ISIN mappings applied")

    await db.execute(delete(AmfiMarketCap))
    if rows_to_insert:
        await db.execute(AmfiMarketCap.__table__.insert(), rows_to_insert)
    await db.flush()

    _write_company_master(rows_to_insert)
    if on_progress:
        await on_progress("Company master CSV updated")

    result = {
        "rows_loaded": len(rows_to_insert),
        "large": counts["Large Cap"],
        "mid": counts["Mid Cap"],
        "small": counts["Small Cap"],
        "sectors_loaded": sectors_loaded,
        "file": xlsx_path.name,
    }
    if file_date:
        result["file_date"] = file_date.strftime("%d %b %Y")
    if stale_warning:
        result["stale_warning"] = stale_warning
    return result


def _resolve_equity_category(
    name: str,
    alias_to_isin: dict[str, str],
    name_to_isin: dict[str, str],
    isin_to_mcap: dict[str, str],
    amfi_by_name: dict[str, str],
) -> str:
    norm = normalize_company_name(name)
    for lookup in (alias_to_isin, name_to_isin):
        isin = lookup.get(norm)
        if isin and isin in isin_to_mcap:
            return isin_to_mcap[isin]
    cat = amfi_by_name.get(norm)
    if cat:
        return cat
    best_r, best_cat = 0.0, None
    for amfi_norm, amfi_cat in amfi_by_name.items():
        r = SequenceMatcher(None, norm, amfi_norm).ratio()
        if r > best_r:
            best_r, best_cat = r, amfi_cat
    if best_r >= 0.85 and best_cat:
        return best_cat
    return "Unclassified Equity"


def _resolve_equity_sector(
    name: str,
    alias_to_isin: dict[str, str],
    name_to_isin: dict[str, str],
    isin_to_sector: dict[str, str],
    name_to_sector: dict[str, str],
) -> str | None:
    norm = normalize_company_name(name)
    for lookup in (alias_to_isin, name_to_isin):
        isin = lookup.get(norm)
        if isin and isin in isin_to_sector:
            return isin_to_sector[isin]
    sec = name_to_sector.get(norm)
    if sec:
        return sec
    best_r, best_s = 0.0, None
    for amfi_norm, amfi_sec in name_to_sector.items():
        r = SequenceMatcher(None, norm, amfi_norm).ratio()
        if r > best_r:
            best_r, best_s = r, amfi_sec
    if best_r >= 0.85:
        return best_s
    return None


_MF_DEBT_RE = re.compile(r"\b(liquid|money\s+market|savings\s+fund|low\s+duration)\b", re.IGNORECASE)
_REIT_RE = re.compile(r"\b(reit|real\s+estate\s+trust)\b", re.IGNORECASE)
# Match "liquid" at a word start (no trailing boundary — catches LIQUIDCASE, LIQUIDBEES, etc.)
_ARBITRAGE_RE = re.compile(r"\barbitrage\b", re.IGNORECASE)

_MV_MULTIPLIERS = {"INR_LAKH": 100_000.0, "INR_CRORE": 10_000_000.0, "INR": 1.0}


async def _fetch_catalog(client: httpx.AsyncClient) -> dict:
    """GET /api/v1/catalog — the full OpenFin catalog keyed by AMFI code."""
    r = await client.get(f"{OPENFIN_BASE}/catalog", timeout=30.0)
    r.raise_for_status()
    return r.json()


async def _fetch_fund_holdings(client: httpx.AsyncClient, amfi_code: str, as_of: str) -> dict | None:
    """GET /api/v1/holdings/{amfi_code}?as_of=... Returns None on any HTTP error."""
    try:
        r = await client.get(f"{OPENFIN_BASE}/holdings/{amfi_code}", params={"as_of": as_of}, timeout=30.0)
        r.raise_for_status()
        return r.json()
    except httpx.HTTPError:
        return None


class _AmfiLookups:
    """Bundles the AMFI-derived lookup tables used for equity classification, so
    they're built once per ingest run and threaded through as a single object."""

    def __init__(self, amfi_all, overrides: dict[str, str], sector_overrides: dict[str, str]):
        self.alias_to_isin: dict[str, str] = {}
        self.name_to_isin: dict[str, str] = {}
        self.isin_to_mcap: dict[str, str] = {}
        self.isin_to_sector: dict[str, str] = {}
        self.amfi_by_name: dict[str, str] = {}
        self.amfi_name_sector: dict[str, str] = {}
        for a in amfi_all:
            self.amfi_by_name[a.name_normalized] = a.categorization
            if a.sector:
                self.amfi_name_sector[a.name_normalized] = a.sector
            if a.isin:
                self.isin_to_mcap[a.isin] = a.categorization
                self.name_to_isin[a.name_normalized] = a.isin
                if a.sector:
                    self.isin_to_sector[a.isin] = a.sector
                if a.aliases:
                    for alias in a.aliases.split("|"):
                        alias_norm = normalize_company_name(alias.strip())
                        if alias_norm:
                            self.alias_to_isin[alias_norm] = a.isin
        self.overrides = overrides
        self.sector_overrides = sector_overrides

    def classify_equity(self, name: str, holding_isin: str | None, fund_isin: str) -> str:
        if holding_isin and not holding_isin.startswith("IN"):
            return "Equity - Foreign"
        if fund_isin in FOREIGN_FUND_ISINS:
            return "Equity - Foreign"
        name_lower = name.lower()
        if any(s in name_lower for s in FOREIGN_COMPANY_SUBSTRINGS):
            return "Equity - Foreign"
        if holding_isin and holding_isin in COMMODITY_ETF_CATEGORY:
            return COMMODITY_ETF_CATEGORY[holding_isin]
        # OpenFin gives us the holding's own ISIN directly — try that exact match
        # before falling back to name-based resolution (which only reaches AMFI's
        # ISIN indirectly, via AMFI's own name wording, and misses whenever the
        # two disclosures spell the company differently).
        if holding_isin and holding_isin in self.isin_to_mcap:
            return self.isin_to_mcap[holding_isin]
        category = _resolve_equity_category(name, self.alias_to_isin, self.name_to_isin, self.isin_to_mcap, self.amfi_by_name)
        if category != "Unclassified Equity":
            return category
        return (
            self.overrides.get(normalize_company_name(name))
            or ETF_CAP_OVERRIDE.get(holding_isin or "")
            or "Unclassified Equity"
        )

    def resolve_sector(self, name: str, holding_isin: str | None = None) -> str | None:
        if holding_isin and holding_isin in self.isin_to_sector:
            return self.isin_to_sector[holding_isin]
        sector = _resolve_equity_sector(name, self.alias_to_isin, self.name_to_isin, self.isin_to_sector, self.amfi_name_sector)
        if sector is None:
            sector = self.sector_overrides.get(normalize_company_name(name))
        return sector


def _classify_non_equity(
    api_holding_type: str,
    section: str,
    instrument_name: str,
    holding_isin: str | None,
    catalog_by_isin: dict[str, dict],
) -> tuple[str, str, str | None]:
    """Returns (stored_holding_type, category, sector) for anything that isn't
    an 'equity' or 'derivative' API holding_type."""
    sec_lower = (section or "").strip().lower()
    text = f"{sec_lower} {instrument_name}".lower()

    if sec_lower in ("certificate of deposit", "commercial paper", "treasury bill"):
        return section.strip(), "Debt", "Fixed Income"
    if api_holding_type == "cash" or sec_lower == "cash":
        return "Cash", "Cash", "Liquid / Money Market"
    if api_holding_type == "fund_unit" or sec_lower == "mutual fund units":
        entry = catalog_by_isin.get(holding_isin) if holding_isin else None
        amfi_category = (entry.get("category") or "").lower() if entry else ""
        if any(k in amfi_category for k in ("debt scheme", "liquid", "money market", "overnight")):
            return "Mutual Fund Units", "Debt", "Liquid / Money Market"
        if "equity scheme" in amfi_category:
            return "Mutual Fund Units", "Other", None
        if _MF_DEBT_RE.search(instrument_name):
            return "Mutual Fund Units", "Debt", "Liquid / Money Market"
        return "Mutual Fund Units", "Other", None
    if api_holding_type in ("commodity", "other"):
        if "gold" in text:
            return "Commodity", "Gold", "Gold"
        if "silver" in text:
            return "Commodity", "Silver", "Silver"
        return "Commodity", "Other", None
    if api_holding_type in ("debt", "money_market"):
        return "Debt", "Debt", "Fixed Income"

    # Unknown API holding_type/section combo — fall through to a sensible default.
    return api_holding_type.capitalize() or "Other", "Other", None


def _map_scheme_holdings(
    fund_isin: str,
    holdings: list[dict],
    mv_multiplier: float,
    catalog_by_isin: dict[str, dict],
    lookups: _AmfiLookups,
    is_arbitrage_fund: bool,
) -> tuple[list[dict], list[dict]]:
    """Maps one fund's raw API holdings into MfSchemeBreakdown-shaped rows
    (holdings_pct not yet set — computed by the caller after totalling).
    Returns (rows, unmatched_equities)."""
    rows: list[dict] = []
    unmatched: list[dict] = []
    plain: list[dict] = []

    equity_by_isin: dict[str, list[dict]] = {}
    deriv_by_isin: dict[str, list[dict]] = {}

    for h in holdings:
        if h.get("market_value") is None:
            continue  # malformed/junk row observed in some disclosures
        # OpenFin sometimes mislabels money-market paper (CDs) as holding_type
        # "equity" — instrument_yield is never set on genuine equity, so it's a
        # reliable tell. Route these to the debt path regardless of fund type.
        is_mislabeled_debt = h["holding_type"] == "equity" and h.get("instrument_yield") is not None
        if is_mislabeled_debt:
            plain.append(h)
        elif is_arbitrage_fund and h["holding_type"] == "equity" and h.get("isin"):
            equity_by_isin.setdefault(h["isin"], []).append(h)
        elif is_arbitrage_fund and h["holding_type"] == "derivative" and h.get("section") == "Futures" and h.get("isin"):
            deriv_by_isin.setdefault(h["isin"], []).append(h)
        else:
            plain.append(h)

    # Arbitrage funds: pair each stock's long equity value against its short futures
    # value (summed across contract expiries) on the same ISIN. The matched (lower)
    # amount is the true arbitrage position; any leftover is unhedged exposure —
    # extra stock if the long side was bigger, a naked short if the derivative was.
    for isin in set(equity_by_isin) | set(deriv_by_isin):
        eq_rows = equity_by_isin.get(isin, [])
        de_rows = deriv_by_isin.get(isin, [])
        long_mv = sum(float(r["market_value"]) for r in eq_rows) * mv_multiplier
        short_mv = -sum(float(r["market_value"]) for r in de_rows) * mv_multiplier
        name = (eq_rows[0] if eq_rows else de_rows[0])["instrument"]
        matched = min(long_mv, short_mv)
        residual = abs(long_mv - short_mv)

        if matched > 0:
            rows.append({
                "name": name, "holding_type": "Arbitrage", "category": "Equity - Arbitrage",
                "sector": lookups.resolve_sector(name, isin), "market_value": matched,
            })
        if residual > 1e-6:
            if long_mv > short_mv:
                category = lookups.classify_equity(name, isin, fund_isin)
                if category == "Unclassified Equity":
                    unmatched.append({"name": name, "scheme_isin": fund_isin})
                rows.append({
                    "name": name, "holding_type": "Equity", "category": category,
                    "sector": lookups.resolve_sector(name, isin), "market_value": residual,
                })
            else:
                rows.append({
                    "name": name, "holding_type": "Derivative", "category": "Derivatives - Leveraged",
                    "sector": lookups.resolve_sector(name, isin), "market_value": -residual,
                })

    for h in plain:
        name = (h.get("instrument") or "").strip()
        if not name:
            continue
        market_value = float(h["market_value"]) * mv_multiplier
        holding_isin = h.get("isin")
        section = h.get("section") or ""

        # OpenFin mislabels some listed equities as holding_type "commodity"/"other"
        # (e.g. Multi Commodity Exchange of India Ltd — the word "Commodity" in the
        # company name apparently confuses their classifier). Real commodity holdings
        # (gold bars, silver, commodity-exchange derivatives) never have an `industry`
        # or a normal equity ISIN — genuine equity always does.
        is_mislabeled_equity = (
            h["holding_type"] in ("commodity", "other")
            and h.get("industry") is not None
            and bool(holding_isin) and holding_isin.startswith("IN")
        )
        api_type = "equity" if is_mislabeled_equity else h["holding_type"]

        if api_type == "equity" and h.get("instrument_yield") is not None:
            holding_type, category, sector = "Certificate of Deposit", "Debt", "Fixed Income"
        elif api_type == "equity":
            if _REIT_RE.search(section) or _REIT_RE.search(name):
                holding_type, category, sector = "Reits", "Real Estate Trust", "Real Estate Trust"
            else:
                category = lookups.classify_equity(name, holding_isin, fund_isin)
                if category == "Unclassified Equity":
                    unmatched.append({"name": name, "scheme_isin": fund_isin})
                holding_type, sector = "Equity", lookups.resolve_sector(name, holding_isin)
        elif api_type == "derivative":
            holding_type, category, sector = "Derivative", "Derivatives - Leveraged", lookups.resolve_sector(name, holding_isin)
        else:
            holding_type, category, sector = _classify_non_equity(api_type, section, name, holding_isin, catalog_by_isin)

        rows.append({
            "name": name, "holding_type": holding_type, "category": category,
            "sector": sector, "market_value": market_value,
        })

    return rows, unmatched


async def ingest_from_openfin(db: AsyncSession, on_progress=None) -> dict:
    resolved = await resolve_scheme_codes(db)
    if on_progress and (resolved["resolved"] or resolved["unresolved"]):
        await on_progress(f"AMFI scheme codes: {resolved['resolved']} resolved, {len(resolved['unresolved'])} unresolved")

    amfi_all = (await db.execute(select(AmfiMarketCap))).scalars().all()
    overrides: dict[str, str] = {
        o.name_normalized: o.category
        for o in (await db.execute(select(EquityCategoryOverride))).scalars().all()
    }
    sector_override_rows = (await db.execute(select(EquitySectorOverride))).scalars().all()
    sector_overrides: dict[str, str] = {o.name_normalized: o.sector for o in sector_override_rows}
    lookups = _AmfiLookups(amfi_all, overrides, sector_overrides)

    # Auto-prune: if AMFI now has a sector for an override, the override is stale
    pruned = 0
    for o in sector_override_rows:
        if _resolve_equity_sector(o.raw_name, lookups.alias_to_isin, lookups.name_to_isin, lookups.isin_to_sector, lookups.amfi_name_sector) is not None:
            await db.execute(delete(EquitySectorOverride).where(EquitySectorOverride.id == o.id))
            sector_overrides.pop(o.name_normalized, None)
            pruned += 1
    if on_progress and pruned:
        await on_progress(f"Auto-removed {pruned} stale sector override(s)")

    held_funds = (await db.execute(
        select(Instrument)
        .join(Holding, Holding.instrument_id == Instrument.id)
        .where(Instrument.instrument_type.in_(("MF", "ETF")))
    )).scalars().all()
    held_isins = {i.isin for i in held_funds if i.isin}
    isin_to_name = {i.isin: (i.tradingsymbol or i.name or i.isin) for i in held_funds if i.isin}

    arbitrage_fund_isins: set[str] = set()
    for i in held_funds:
        name_to_check = " ".join(filter(None, [i.tradingsymbol, i.name]))
        if i.isin and _ARBITRAGE_RE.search(name_to_check):
            arbitrage_fund_isins.add(i.isin)

    schemes_processed = 0
    schemes_skipped = 0
    rows_upserted = 0
    unmatched: list[dict] = []
    errors: list[str] = []
    seen_isins: set[str] = set()
    latest_as_of: date | None = None

    async with httpx.AsyncClient(follow_redirects=True, headers={"User-Agent": "portfolio-mac-arm/1.0"}) as client:
        try:
            catalog = await _fetch_catalog(client)
        except httpx.HTTPError as e:
            return {"schemes_processed": 0, "rows_upserted": 0, "unmatched_equities": [],
                    "missing_funds": [], "errors": [f"OpenFin catalog fetch failed: {e}"]}

        catalog_by_isin = {e["isin"]: e for e in catalog.values() if e.get("isin")}

        # Resolve each held fund to a catalog entry: by amfi_scheme_code first, then by ISIN.
        fund_catalog_entry: dict[str, dict] = {}  # scheme_isin -> catalog entry
        for i in held_funds:
            if not i.isin:
                continue
            entry = catalog.get(i.amfi_scheme_code or "") or catalog_by_isin.get(i.isin)
            if entry and entry.get("has_holdings", True):
                fund_catalog_entry[i.isin] = entry

        missing_funds = [
            {"isin": isin, "name": isin_to_name.get(isin, isin)}
            for isin in sorted(held_isins - set(fund_catalog_entry))
        ]

        # Staleness check: compare the catalog's latest_as_of against what we have
        # stored locally per scheme. Only schemes with a newer disclosure are fetched.
        local_as_of_rows = (await db.execute(
            select(MfSchemeBreakdown.scheme_isin, MfSchemeBreakdown.as_of).distinct()
        )).all()
        local_as_of = {isin: d for isin, d in local_as_of_rows if d is not None}

        stale: list[tuple[str, dict]] = []
        for isin, entry in fund_catalog_entry.items():
            catalog_as_of = date.fromisoformat(entry["latest_as_of"])
            if isin not in local_as_of or catalog_as_of > local_as_of[isin]:
                stale.append((isin, entry))
            else:
                schemes_skipped += 1
                seen_isins.add(isin)
                if latest_as_of is None or local_as_of[isin] > latest_as_of:
                    latest_as_of = local_as_of[isin]

        if not stale:
            return {
                "schemes_processed": 0, "rows_upserted": 0, "schemes_skipped": schemes_skipped,
                "already_current": True, "as_of": latest_as_of.isoformat() if latest_as_of else None,
                "unmatched_equities": [], "missing_funds": missing_funds, "errors": [],
            }

        if on_progress:
            await on_progress(f"{len(stale)} scheme(s) have newer disclosures ({schemes_skipped} already current)")

        for isin, entry in stale:
            fund_name = isin_to_name.get(isin, isin)
            as_of_str = entry["latest_as_of"]
            if on_progress:
                await on_progress(f"[{schemes_processed + 1}/{len(stale)}] {fund_name} (as of {as_of_str})")

            response = await _fetch_fund_holdings(client, entry["amfi_code"], as_of_str)
            if response is None:
                errors.append(f"{fund_name}: holdings fetch failed")
                continue

            mv_unit = response.get("meta", {}).get("market_value_unit")
            if mv_unit not in _MV_MULTIPLIERS:
                errors.append(f"{fund_name}: unrecognised market_value_unit {mv_unit!r} — skipped")
                continue

            rows, scheme_unmatched = _map_scheme_holdings(
                isin, response.get("holdings", []), _MV_MULTIPLIERS[mv_unit],
                catalog_by_isin, lookups, isin in arbitrage_fund_isins,
            )
            unmatched.extend(scheme_unmatched)

            total_mv = sum(r["market_value"] for r in rows)
            scheme_as_of = date.fromisoformat(response["meta"]["as_of"])
            values = [{
                "scheme_isin": isin,
                "name": r["name"][:255],
                "holding_type": r["holding_type"][:50],
                "holdings_pct": round(r["market_value"] / total_mv * 100, 8) if total_mv else 0.0,
                "market_value": round(r["market_value"], 2),
                "category": r["category"],
                "sector": r["sector"],
                "as_of": scheme_as_of,
                "updated_at": now_ist(),
            } for r in rows]

            # Replace, not upsert: this scheme's entire local breakdown is wiped
            # and rebuilt from the fresh disclosure, not merged row-by-row.
            await db.execute(delete(MfSchemeBreakdown).where(MfSchemeBreakdown.scheme_isin == isin))
            if values:
                await db.execute(MfSchemeBreakdown.__table__.insert(), values)
                rows_upserted += len(values)

            seen_isins.add(isin)
            schemes_processed += 1
            if latest_as_of is None or scheme_as_of > latest_as_of:
                latest_as_of = scheme_as_of
            if on_progress:
                unmatched_here = sum(1 for u in scheme_unmatched if u["scheme_isin"] == isin)
                msg = f"  → {len(values)} rows"
                if unmatched_here:
                    msg += f", {unmatched_here} unmatched"
                await on_progress(msg)

    # Synthesize 100% rows for commodity ETFs that OpenFin doesn't cover at all.
    for isin, commodity_cat in COMMODITY_ETF_CATEGORY.items():
        if isin not in held_isins or isin in seen_isins:
            continue
        await db.execute(delete(MfSchemeBreakdown).where(MfSchemeBreakdown.scheme_isin == isin))
        await db.execute(MfSchemeBreakdown.__table__.insert(), [{
            "scheme_isin": isin,
            "name": commodity_cat,
            "holding_type": commodity_cat,
            "holdings_pct": 100.0,
            "category": commodity_cat,
            "sector": commodity_cat,
            "updated_at": now_ist(),
        }])
        seen_isins.add(isin)
        rows_upserted += 1
        if on_progress:
            await on_progress(f"  → {isin_to_name.get(isin, isin)}: synthesized 100% {commodity_cat} (not in OpenFin)")

    # Remove rows for held schemes that dropped out of both the stale-fetch set and catalog.
    stale_isins = {isin for isin, _ in stale}
    to_clean = held_isins - seen_isins - (set(fund_catalog_entry) - stale_isins)
    if to_clean:
        await db.execute(delete(MfSchemeBreakdown).where(MfSchemeBreakdown.scheme_isin.in_(to_clean)))

    await db.commit()

    return {
        "schemes_processed": schemes_processed,
        "rows_upserted": rows_upserted,
        "schemes_skipped": schemes_skipped,
        "already_current": False,
        "as_of": latest_as_of.isoformat() if latest_as_of else None,
        "unmatched_equities": unmatched,
        "missing_funds": missing_funds,
        "errors": errors[:30],
    }
